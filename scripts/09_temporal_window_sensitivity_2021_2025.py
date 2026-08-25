"""Fixed-sample temporal-window sensitivity analysis for visitor demand.

This standalone script compares the manuscript's primary 2023--2025 visitor-
demand window with an extended 2021--2025 window for the same destinations.
It does not depend on a notebook state, Google Drive, upload widgets, or an
embedded copy of the destination data.

Required inputs
---------------
1. Official visitor workbook for 2021.
2. Official visitor workbook for 2022.
3. Frozen destination-demand workbook/CSV produced by the primary analysis.

Important analytical rules
--------------------------
* The destination sample is fixed by the frozen primary-demand input.
* Missing destination-year values remain missing and are never replaced by 0.
* Mean demand is calculated from the available annual observations.
* Historical destinations without frozen AECS metadata are audited but excluded.
* The generic 2022 label ``Embung`` is treated as ambiguous and is not linked to
  ``Embung Paccekke``.
* Adding years increases temporal coverage, not the number of independent
  destination-level observations.

Example
-------
python scripts/09_temporal_window_sensitivity_2021_2025.py \
  --visitor-2021 "/path/to/Permintaan data 31 Des 2021.xlsx" \
  --visitor-2022 "/path/to/Permintaan data 30 Des 2022.xlsx" \
  --primary-demand "/path/to/FINAL_HYBRID_AECS_VISITOR_DEMAND_RESULTS.xlsx" \
  --output-dir "outputs/09_temporal_window_sensitivity_2021_2025"
"""

from __future__ import annotations

import argparse
import itertools
import re
import sys
import unicodedata
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import spearmanr


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent

DEFAULT_OUTPUT_DIR = (
    REPO_ROOT / "outputs" / "09_temporal_window_sensitivity_2021_2025"
)
DEFAULT_SEED = 20260825
DEFAULT_N_PERM = 250_000
DEFAULT_N_BOOT = 10_000
DEFAULT_ALPHA = 0.05
CLASS_THRESHOLD = 2 / 3

MONTH_NAMES = {
    "januari",
    "februari",
    "maret",
    "april",
    "mei",
    "juni",
    "juli",
    "agustus",
    "september",
    "oktober",
    "november",
    "desember",
}

DESTINATION_ALIASES = {
    "embun paccekke": "Embung Paccekke",
    "embung paccekke": "Embung Paccekke",
    "ujung batu": "Pantai Ujung Batu",
    "pantai ujung batu": "Pantai Ujung Batu",
    "diana waterpark": "Diana Water Park",
    "diana water park": "Diana Water Park",
    "pekkae ecolodge": "PekkaE Ecolodge",
}


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------


def normalize_text(value: object) -> str:
    """Return a conservative comparison key for names and column labels."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower().replace("’", "'").replace("`", "'")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_key(value: object) -> str:
    return normalize_text(value).replace(" ", "")


def numeric(value: object) -> float:
    if value is None or value == "":
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return np.nan


def find_col(
    df: pd.DataFrame,
    candidates: Iterable[str],
    *,
    required: bool = True,
) -> str | None:
    lookup = {clean_key(column): column for column in df.columns}
    for candidate in candidates:
        match = lookup.get(clean_key(candidate))
        if match is not None:
            return match
    if required:
        raise KeyError(
            "Required column not found. "
            f"Candidates={list(candidates)}; available={list(df.columns)}"
        )
    return None


def sample_midrank_percentile(values: pd.Series) -> pd.Series:
    n = int(values.notna().sum())
    if n == 0:
        return pd.Series(np.nan, index=values.index)
    return (values.rank(method="average", ascending=True) - 0.5) / n


def mismatch_class(
    supply: np.ndarray,
    demand: np.ndarray,
    threshold: float = CLASS_THRESHOLD,
) -> np.ndarray:
    supply_high = supply >= threshold
    demand_high = demand >= threshold
    result = np.full(len(supply), "Lower readiness–lower demand", dtype=object)
    result[(~supply_high) & demand_high] = "Lower readiness–high demand"
    result[supply_high & (~demand_high)] = "High readiness–lower demand"
    result[supply_high & demand_high] = "High readiness–high demand"
    return result


# ---------------------------------------------------------------------------
# Official historical-workbook reader
# ---------------------------------------------------------------------------


def sheet_score(ws, target_year: int) -> int:
    """Score a worksheet using its content, not merely its sheet name."""
    sample_values: list[str] = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, 45),
        min_col=1,
        max_col=min(ws.max_column, 40),
        values_only=True,
    ):
        sample_values.extend(normalize_text(value) for value in row if value is not None)
    content = " | ".join(sample_values)
    score = 0
    if re.search(rf"\btahun\s+{target_year}\b", content):
        score += 100
    elif re.search(rf"\b{target_year}\b", content):
        score += 25
    if "obyek wisata" in content or "objek wisata" in content:
        score += 20
    if "wisatawan nusantara" in content:
        score += 10
    if "wisatawan mancanegara" in content:
        score += 10
    if str(target_year) in normalize_text(ws.title):
        score += 5
    return score


def select_year_sheet(workbook_path: Path, target_year: int):
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    scored = [(sheet_score(ws, target_year), ws.title, ws) for ws in workbook.worksheets]
    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_title, best_sheet = scored[0]
    if best_score < 100:
        details = ", ".join(f"{title}={score}" for score, title, _ in scored)
        raise ValueError(
            f"No worksheet explicitly identified as year {target_year} in "
            f"{workbook_path}. Sheet scores: {details}"
        )
    return workbook, best_sheet, best_title, best_score


def find_header_row(ws, start_row: int, end_row: int) -> int:
    for row_idx in range(start_row, min(end_row, ws.max_row) + 1):
        values = [normalize_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        if any(value in {"obyek wisata", "objek wisata"} for value in values):
            return row_idx
    raise ValueError(
        f"Destination-name header was not found between rows {start_row} and {end_row}."
    )


def section_rows(ws) -> list[tuple[int, str]]:
    sections: list[tuple[int, str]] = []
    for row_idx in range(1, ws.max_row + 1):
        row_text = " ".join(
            normalize_text(ws.cell(row_idx, col).value)
            for col in range(1, min(ws.max_column, 15) + 1)
        )
        if "wisatawan nusantara" in row_text:
            sections.append((row_idx, "domestic"))
        elif "wisatawan mancanegara" in row_text:
            sections.append((row_idx, "foreign"))
    if not sections:
        raise ValueError("Domestic/foreign visitor sections were not found.")
    return sections


def header_columns(ws, header_row: int) -> tuple[int, int, list[list[int]]]:
    headers = [normalize_text(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    name_candidates = [index + 1 for index, value in enumerate(headers) if value in {"obyek wisata", "objek wisata"}]
    if not name_candidates:
        raise ValueError(f"No destination-name column in header row {header_row}.")
    name_col = name_candidates[0]

    total_candidates = [
        index + 1
        for index, value in enumerate(headers)
        if value in {"jumlah", "total"} and index + 1 > name_col
    ]
    if not total_candidates:
        raise ValueError(f"No annual-total column in header row {header_row}.")
    total_col = total_candidates[-1]

    month_starts = [
        index + 1
        for index, value in enumerate(headers)
        if value in MONTH_NAMES and name_col < index + 1 < total_col
    ]
    month_groups: list[list[int]] = []
    for idx, start in enumerate(month_starts):
        stop = month_starts[idx + 1] if idx + 1 < len(month_starts) else total_col
        month_groups.append(list(range(start, stop)))
    return name_col, total_col, month_groups


def parse_visitor_section(
    ws,
    section_start: int,
    section_end: int,
    visitor_type: str,
) -> list[dict[str, object]]:
    header_row = find_header_row(ws, section_start, min(section_start + 12, section_end))
    name_col, total_col, month_groups = header_columns(ws, header_row)
    rows: list[dict[str, object]] = []

    for row_idx in range(header_row + 1, section_end + 1):
        serial = numeric(ws.cell(row_idx, 1).value)
        original_name = ws.cell(row_idx, name_col).value
        if not np.isfinite(serial) or not normalize_text(original_name):
            continue
        total = numeric(ws.cell(row_idx, total_col).value)
        if not np.isfinite(total):
            continue

        month_observed: list[bool] = []
        for group in month_groups:
            group_values = [numeric(ws.cell(row_idx, col).value) for col in group]
            month_observed.append(any(np.isfinite(value) for value in group_values))
        complete = len(month_observed) == 12 and all(month_observed)

        rows.append(
            {
                "destination_original": str(original_name).strip(),
                "visitor_type": visitor_type,
                "visitors": int(round(total)),
                "complete_12_months": bool(complete),
            }
        )
    return rows


def read_historical_workbook(workbook_path: Path, target_year: int) -> tuple[pd.DataFrame, dict[str, object]]:
    workbook, ws, sheet_name, score = select_year_sheet(workbook_path, target_year)
    sections = section_rows(ws)
    parsed: list[dict[str, object]] = []
    for index, (start, label) in enumerate(sections):
        end = sections[index + 1][0] - 1 if index + 1 < len(sections) else ws.max_row
        parsed.extend(parse_visitor_section(ws, start, end, label))
    workbook.close()

    raw = pd.DataFrame(parsed)
    if raw.empty:
        raise ValueError(f"No visitor records were parsed from {workbook_path}::{sheet_name}.")
    raw["normalized_name"] = raw["destination_original"].map(normalize_text)
    grouped = (
        raw.groupby("normalized_name", as_index=False)
        .agg(
            destination_original=("destination_original", "first"),
            visitors=("visitors", "sum"),
            complete_12_months=("complete_12_months", "max"),
        )
    )
    grouped["year"] = target_year
    grouped["source_workbook"] = str(workbook_path.resolve())
    grouped["source_sheet"] = sheet_name
    audit = {
        "year": target_year,
        "source_workbook": str(workbook_path.resolve()),
        "source_sheet": sheet_name,
        "sheet_selection_score": score,
        "n_destinations": int(len(grouped)),
        "total_reported_visitors": int(grouped["visitors"].sum()),
        "complete_12_month_rows": int(grouped["complete_12_months"].sum()),
    }
    return grouped, audit


# ---------------------------------------------------------------------------
# Frozen primary-demand reader and name reconciliation
# ---------------------------------------------------------------------------


def score_primary_frame(frame: pd.DataFrame) -> int:
    required_groups = [
        ["destination", "destination_name", "nama_destinasi"],
        ["AECS", "AECS_final"],
        ["supply_percentile", "readiness_percentile"],
        ["visitors_2023", "2023"],
        ["visitors_2024", "2024"],
        ["visitors_2025", "2025"],
    ]
    return sum(
        find_col(frame, candidates, required=False) is not None
        for candidates in required_groups
    )


def read_primary_source(path: Path) -> tuple[pd.DataFrame, str]:
    if path.suffix.lower() == ".csv":
        frame = pd.read_csv(path)
        if score_primary_frame(frame) < 6:
            raise ValueError(f"CSV does not contain the required primary-demand fields: {path}")
        return frame, str(path.resolve())

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Primary demand must be a .csv, .xlsx, or .xlsm file.")
    book = pd.ExcelFile(path)
    candidates: list[tuple[int, str, pd.DataFrame]] = []
    for sheet in book.sheet_names:
        frame = pd.read_excel(path, sheet_name=sheet)
        candidates.append((score_primary_frame(frame), sheet, frame))
    candidates.sort(key=lambda item: item[0], reverse=True)
    best_score, best_sheet, best_frame = candidates[0]
    if best_score < 6:
        details = ", ".join(f"{sheet}={score}" for score, sheet, _ in candidates)
        raise ValueError(
            f"No sheet in {path} has all required primary-demand fields. Scores: {details}"
        )
    return best_frame, f"{path.resolve()}::{best_sheet}"


def standardize_primary(frame: pd.DataFrame) -> pd.DataFrame:
    rename: dict[str, str] = {}
    aliases = {
        "destination": ["destination", "destination_name", "nama_destinasi"],
        "AECS": ["AECS", "AECS_final"],
        "supply_percentile": ["supply_percentile", "readiness_percentile"],
        "visitors_2023": ["visitors_2023", "2023"],
        "visitors_2024": ["visitors_2024", "2024"],
        "visitors_2025": ["visitors_2025", "2025"],
        "coordinate_status": ["coordinate_status"],
        "spatial_assignment_method": ["spatial_assignment_method"],
    }
    for target, options in aliases.items():
        source = find_col(frame, options, required=target not in {"coordinate_status", "spatial_assignment_method"})
        if source is not None and source != target:
            rename[source] = target
    result = frame.rename(columns=rename).copy()
    for optional in ("coordinate_status", "spatial_assignment_method"):
        if optional not in result:
            result[optional] = pd.NA

    keep = [
        "destination",
        "visitors_2023",
        "visitors_2024",
        "visitors_2025",
        "AECS",
        "supply_percentile",
        "coordinate_status",
        "spatial_assignment_method",
    ]
    result = result[keep].copy()
    result["destination"] = result["destination"].astype(str).str.strip()
    for column in ["visitors_2023", "visitors_2024", "visitors_2025", "AECS", "supply_percentile"]:
        result[column] = pd.to_numeric(result[column], errors="coerce")

    if result["destination"].duplicated().any():
        duplicates = result.loc[result["destination"].duplicated(False), "destination"].tolist()
        raise ValueError(f"Duplicate destinations in frozen primary input: {duplicates}")
    if result[["AECS", "supply_percentile"]].isna().any().any():
        raise ValueError("Every primary destination must have AECS and supply_percentile values.")
    observed = result[["visitors_2023", "visitors_2024", "visitors_2025"]].notna().sum(axis=1)
    result = result.loc[observed >= 2].reset_index(drop=True)
    if result.empty:
        raise ValueError("No destination has at least two observations in 2023--2025.")
    return result


def reconcile_historical_names(historical: pd.DataFrame, primary: pd.DataFrame) -> pd.DataFrame:
    primary_lookup = {normalize_text(name): name for name in primary["destination"]}
    rows: list[dict[str, object]] = []
    for record in historical.to_dict("records"):
        original = record["destination_original"]
        key = normalize_text(original)
        alias = DESTINATION_ALIASES.get(key)
        canonical_key = normalize_text(alias) if alias else key

        if key == "embung":
            destination = pd.NA
            status = "ambiguous_excluded"
        elif canonical_key in primary_lookup:
            destination = primary_lookup[canonical_key]
            status = "matched_primary"
        else:
            destination = alias if alias else original
            status = "confirmed_unmatched"

        rows.append(
            {
                **record,
                "destination": destination,
                "name_match_status": status,
            }
        )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Window construction and association diagnostics
# ---------------------------------------------------------------------------


def calculate_window(frame: pd.DataFrame, years: list[int], label: str) -> pd.DataFrame:
    result = frame.copy()
    annual_columns = [f"visitors_{year}" for year in years]
    result[f"observed_years_{years[0]}_{years[-1]}"] = result[annual_columns].notna().sum(axis=1)
    mean_col = f"mean_annual_visitors_{years[0]}_{years[-1]}"
    rank_col = f"demand_rank_{years[0]}_{years[-1]}"
    percentile_col = f"demand_percentile_{years[0]}_{years[-1]}"
    mismatch_col = f"mismatch_{years[0]}_{years[-1]}"
    class_col = f"supply_demand_class_{years[0]}_{years[-1]}"

    result[mean_col] = result[annual_columns].mean(axis=1, skipna=True)
    result[rank_col] = result[mean_col].rank(method="average", ascending=True)
    result[percentile_col] = sample_midrank_percentile(result[mean_col])
    result[mismatch_col] = result[percentile_col] - result["supply_percentile"]
    result[class_col] = mismatch_class(
        result["supply_percentile"].to_numpy(float),
        result[percentile_col].to_numpy(float),
    )
    result["analysis_window"] = label
    result["mean_annual_visitors"] = result[mean_col]
    result["demand_percentile"] = result[percentile_col]
    return result


def paired_bootstrap_spearman(
    x: np.ndarray,
    y: np.ndarray,
    n_boot: int,
    seed: int,
) -> tuple[float, float]:
    rng = np.random.default_rng(seed)
    values: list[float] = []
    n = len(x)
    for _ in range(n_boot):
        index = rng.integers(0, n, n)
        x_boot, y_boot = x[index], y[index]
        if np.unique(x_boot).size < 2 or np.unique(y_boot).size < 2:
            continue
        rho = float(spearmanr(x_boot, y_boot).statistic)
        if np.isfinite(rho):
            values.append(rho)
    if not values:
        return np.nan, np.nan
    low, high = np.quantile(values, [0.025, 0.975])
    return float(low), float(high)


def rank_corr_against_permutations(x: np.ndarray, permuted_y: np.ndarray) -> np.ndarray:
    rank_x = np.argsort(np.argsort(x)).astype(float)
    rank_x -= rank_x.mean()
    rank_y = np.argsort(np.argsort(permuted_y, axis=1), axis=1).astype(float)
    rank_y -= rank_y.mean(axis=1, keepdims=True)
    return (rank_y * rank_x).sum(axis=1) / np.sqrt(
        (rank_x * rank_x).sum() * (rank_y * rank_y).sum(axis=1)
    )


def permutation_pvalue(x: np.ndarray, y: np.ndarray, n_perm: int, seed: int) -> float:
    observed = abs(float(spearmanr(x, y).statistic))
    n = len(x)
    if n <= 8:
        permuted = np.asarray(list(itertools.permutations(y)), dtype=float)
        null_rho = rank_corr_against_permutations(x, permuted)
        return float(np.mean(np.abs(null_rho) >= observed - 1e-12))
    rng = np.random.default_rng(seed)
    order = np.argsort(rng.random((n_perm, n)), axis=1)
    null_rho = rank_corr_against_permutations(x, y[order])
    return float((np.sum(np.abs(null_rho) >= observed - 1e-12) + 1) / (n_perm + 1))


def critical_spearman(n: int, alpha: float, n_perm: int, seed: int) -> tuple[float, float]:
    base = np.arange(n, dtype=float)
    if n <= 8:
        permuted = np.asarray(list(itertools.permutations(base)), dtype=float)
    else:
        rng = np.random.default_rng(seed)
        permuted = np.argsort(rng.random((n_perm, n)), axis=1).astype(float)
    absolute_null = np.abs(rank_corr_against_permutations(base, permuted))
    for critical in np.unique(np.round(absolute_null, 12)):
        tail = float(np.mean(absolute_null >= critical - 1e-12))
        if tail <= alpha:
            return float(critical), tail
    return 1.0, float(np.mean(absolute_null >= 1.0 - 1e-12))


def available_samples(frame: pd.DataFrame) -> dict[str, pd.DataFrame]:
    samples = {"Primary": frame}
    coordinate_available = frame["coordinate_status"].notna().any()
    assignment_available = frame["spatial_assignment_method"].notna().any()
    if coordinate_available:
        samples["Non-manual"] = frame[
            frame["coordinate_status"].astype(str).str.lower() != "manual_approximate"
        ]
    if coordinate_available and assignment_available:
        samples["Strict point-within-grid"] = frame[
            (frame["coordinate_status"].astype(str).str.lower() != "manual_approximate")
            & (
                frame["spatial_assignment_method"].astype(str).str.lower()
                == "point_within_grid"
            )
        ]
    return samples


def association_audit(
    windows: list[tuple[str, pd.DataFrame]],
    *,
    n_perm: int,
    n_boot: int,
    alpha: float,
    seed: int,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for window_index, (window_label, frame) in enumerate(windows):
        for sample_index, (sample_label, sample) in enumerate(available_samples(frame).items()):
            if len(sample) < 4:
                continue
            x = sample["AECS"].to_numpy(float)
            y = sample["mean_annual_visitors"].to_numpy(float)
            # A ten-unit window offset preserves the deterministic seeds used
            # for the manuscript's temporal-window sensitivity calculation.
            bootstrap_seed = seed + window_index * 10 + sample_index
            permutation_seed = seed + 100 + window_index * 10 + sample_index
            critical_seed = seed + 200 + sample_index
            rho, asymptotic_p = spearmanr(x, y)
            ci_low, ci_high = paired_bootstrap_spearman(x, y, n_boot, bootstrap_seed)
            permutation_p = permutation_pvalue(x, y, n_perm, permutation_seed)
            critical, actual_tail = critical_spearman(
                len(sample), alpha, n_perm, critical_seed
            )
            rows.append(
                {
                    "analysis_window": window_label,
                    "sample": sample_label,
                    "n": len(sample),
                    "rho": float(rho),
                    "asymptotic_p": float(asymptotic_p),
                    "permutation_p": permutation_p,
                    "bootstrap_95CI_low": ci_low,
                    "bootstrap_95CI_high": ci_high,
                    "critical_abs_rho_alpha_0.05": critical,
                    "actual_null_tail_at_critical": actual_tail,
                    "bootstrap_seed": bootstrap_seed,
                    "permutation_seed": permutation_seed,
                    "n_bootstrap": n_boot,
                    "n_permutations": n_perm,
                }
            )
    return pd.DataFrame(rows)


def compare_windows(primary: pd.DataFrame, extended: pd.DataFrame) -> pd.DataFrame:
    primary_suffix = "2023_2025"
    extended_suffix = "2021_2025"
    primary_columns = [
        "destination",
        f"observed_years_{primary_suffix}",
        f"mean_annual_visitors_{primary_suffix}",
        f"demand_rank_{primary_suffix}",
        f"demand_percentile_{primary_suffix}",
        f"mismatch_{primary_suffix}",
        f"supply_demand_class_{primary_suffix}",
    ]
    extended_columns = [
        "destination",
        f"observed_years_{extended_suffix}",
        f"mean_annual_visitors_{extended_suffix}",
        f"demand_rank_{extended_suffix}",
        f"demand_percentile_{extended_suffix}",
        f"mismatch_{extended_suffix}",
        f"supply_demand_class_{extended_suffix}",
    ]
    comparison = primary[primary_columns].merge(
        extended[extended_columns], on="destination", how="inner", validate="one_to_one"
    )
    comparison["rank_change_extended_minus_primary"] = (
        comparison[f"demand_rank_{extended_suffix}"]
        - comparison[f"demand_rank_{primary_suffix}"]
    )
    comparison["mismatch_change_extended_minus_primary"] = (
        comparison[f"mismatch_{extended_suffix}"]
        - comparison[f"mismatch_{primary_suffix}"]
    )
    comparison["class_retained"] = (
        comparison[f"supply_demand_class_{extended_suffix}"]
        == comparison[f"supply_demand_class_{primary_suffix}"]
    )
    return comparison.sort_values(
        ["class_retained", "destination"], ascending=[True, True]
    ).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Input discovery, output tables, and command-line entry point
# ---------------------------------------------------------------------------


def discover_file(explicit: str | None, patterns: list[str], label: str) -> Path:
    if explicit:
        path = Path(explicit).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"{label} not found: {path}")
        return path
    matches: list[Path] = []
    for pattern in patterns:
        matches.extend(REPO_ROOT.glob(pattern))
    matches = sorted({path.resolve() for path in matches if path.is_file()})
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise FileNotFoundError(
            f"{label} was not found automatically. Provide its path explicitly."
        )
    raise FileNotFoundError(
        f"Multiple candidates were found for {label}; choose one explicitly:\n- "
        + "\n- ".join(str(path) for path in matches)
    )


def parser() -> argparse.ArgumentParser:
    argument_parser = argparse.ArgumentParser(
        description="Compare fixed-sample visitor-demand windows 2023--2025 and 2021--2025."
    )
    argument_parser.add_argument("--visitor-2021", help="Official 2021 visitor workbook.")
    argument_parser.add_argument("--visitor-2022", help="Official 2022 visitor workbook.")
    argument_parser.add_argument(
        "--primary-demand",
        help="Frozen primary destination-demand workbook or CSV containing 2023--2025.",
    )
    argument_parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help=f"Output directory (default: {DEFAULT_OUTPUT_DIR}).",
    )
    argument_parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    argument_parser.add_argument("--n-perm", type=int, default=DEFAULT_N_PERM)
    argument_parser.add_argument("--n-boot", type=int, default=DEFAULT_N_BOOT)
    argument_parser.add_argument("--alpha", type=float, default=DEFAULT_ALPHA)
    argument_parser.add_argument(
        "--expected-n",
        type=int,
        default=12,
        help="Expected fixed destination sample; set 0 to disable the check.",
    )
    return argument_parser


def write_outputs(
    output_dir: Path,
    input_audit: pd.DataFrame,
    historical_matches: pd.DataFrame,
    historical_long: pd.DataFrame,
    primary: pd.DataFrame,
    extended: pd.DataFrame,
    associations: pd.DataFrame,
    comparison: pd.DataFrame,
    primary_source: str,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    input_audit.to_csv(output_dir / "01_historical_input_audit.csv", index=False)
    historical_matches.to_csv(
        output_dir / "02_historical_destination_matches.csv", index=False
    )
    historical_long.to_csv(output_dir / "03_historical_long.csv", index=False)
    primary.to_csv(output_dir / "04_primary_2023_2025_standardized.csv", index=False)
    extended.to_csv(output_dir / "05_extended_2021_2025_standardized.csv", index=False)
    associations.to_csv(output_dir / "06_association_windows.csv", index=False)

    panel_a = associations[associations["sample"] == "Primary"].copy()
    panel_a["window"] = panel_a["analysis_window"]
    panel_a = panel_a[
        [
            "window",
            "n",
            "rho",
            "permutation_p",
            "bootstrap_95CI_low",
            "bootstrap_95CI_high",
            "critical_abs_rho_alpha_0.05",
        ]
    ]
    panel_a.to_csv(output_dir / "07_supplementary_table_s8_panel_a.csv", index=False)

    panel_b = comparison[
        [
            "destination",
            "demand_rank_2023_2025",
            "demand_rank_2021_2025",
            "mismatch_2023_2025",
            "mismatch_2021_2025",
            "supply_demand_class_2023_2025",
            "supply_demand_class_2021_2025",
            "class_retained",
        ]
    ].copy()
    panel_b.to_csv(output_dir / "08_supplementary_table_s8_panel_b.csv", index=False)
    comparison.to_csv(output_dir / "09_window_comparison.csv", index=False)

    primary_result = panel_a.loc[panel_a["window"] == "Primary 2023–2025"].iloc[0]
    extended_result = panel_a.loc[panel_a["window"] == "Extended 2021–2025"].iloc[0]
    retained = int(comparison["class_retained"].sum())
    changed = comparison.loc[~comparison["class_retained"], "destination"].tolist()
    summary = pd.DataFrame(
        [
            ("Primary source", primary_source),
            ("Primary sample size", int(primary_result["n"])),
            ("Extended sample size", int(extended_result["n"])),
            ("Primary rho", float(primary_result["rho"])),
            ("Extended rho", float(extended_result["rho"])),
            ("Primary permutation p", float(primary_result["permutation_p"])),
            ("Extended permutation p", float(extended_result["permutation_p"])),
            ("Classes retained", f"{retained} of {len(comparison)}"),
            ("Class-retention percentage", 100 * retained / len(comparison)),
            ("Changed destinations", ", ".join(changed) if changed else "None"),
        ],
        columns=["diagnostic", "value"],
    )
    summary.to_csv(output_dir / "10_temporal_window_summary.csv", index=False)

    summary_text = f"""TEMPORAL-WINDOW SENSITIVITY COMPLETED
=====================================
Primary source: {primary_source}

Primary 2023–2025: n={int(primary_result['n'])}; rho={primary_result['rho']:.3f};
permutation p={primary_result['permutation_p']:.3f}; bootstrap 95% CI
[{primary_result['bootstrap_95CI_low']:.3f}, {primary_result['bootstrap_95CI_high']:.3f}].

Extended 2021–2025: n={int(extended_result['n'])}; rho={extended_result['rho']:.3f};
permutation p={extended_result['permutation_p']:.3f}; bootstrap 95% CI
[{extended_result['bootstrap_95CI_low']:.3f}, {extended_result['bootstrap_95CI_high']:.3f}].

Class retained: {retained} of {len(comparison)} destinations
({100 * retained / len(comparison):.1f}%).
Changed destinations: {', '.join(changed) if changed else 'None'}.

Interpretation: extending the temporal window increases temporal coverage but
does not increase destination-level sample size or resolve limited statistical
power. Class changes should be interpreted as threshold sensitivity.
"""
    (output_dir / "RUN_SUMMARY.txt").write_text(summary_text, encoding="utf-8")


def main() -> int:
    args = parser().parse_args()
    visitor_2021 = discover_file(
        args.visitor_2021,
        ["data/raw/**/*2021*.xlsx", "data/private/**/*2021*.xlsx"],
        "2021 visitor workbook",
    )
    visitor_2022 = discover_file(
        args.visitor_2022,
        ["data/raw/**/*2022*.xlsx", "data/private/**/*2022*.xlsx"],
        "2022 visitor workbook",
    )
    primary_path = discover_file(
        args.primary_demand,
        [
            "outputs/**/FINAL_HYBRID_AECS_VISITOR_DEMAND_RESULTS.xlsx",
            "data/processed/**/*demand*.csv",
        ],
        "frozen primary-demand file",
    )
    output_dir = Path(args.output_dir).expanduser().resolve()

    historical_2021, audit_2021 = read_historical_workbook(visitor_2021, 2021)
    historical_2022, audit_2022 = read_historical_workbook(visitor_2022, 2022)
    input_audit = pd.DataFrame([audit_2021, audit_2022])

    primary_raw, primary_source = read_primary_source(primary_path)
    primary_base = standardize_primary(primary_raw)
    if args.expected_n and len(primary_base) != args.expected_n:
        raise ValueError(
            f"Expected n={args.expected_n} primary destinations, but found "
            f"n={len(primary_base)} after the >=2-year rule."
        )

    historical = pd.concat([historical_2021, historical_2022], ignore_index=True)
    historical_matches = reconcile_historical_names(historical, primary_base)
    matched = historical_matches[
        historical_matches["name_match_status"] == "matched_primary"
    ].copy()
    historical_wide = matched.pivot_table(
        index="destination", columns="year", values="visitors", aggfunc="sum"
    ).rename(columns={2021: "visitors_2021", 2022: "visitors_2022"})

    base = primary_base.merge(
        historical_wide, left_on="destination", right_index=True, how="left"
    )
    for column in ("visitors_2021", "visitors_2022"):
        if column not in base:
            base[column] = np.nan
    ordered_years = [f"visitors_{year}" for year in range(2021, 2026)]
    base[ordered_years] = base[ordered_years].apply(pd.to_numeric, errors="coerce")

    primary_window = calculate_window(base, [2023, 2024, 2025], "Primary 2023–2025")
    extended_window = calculate_window(base, [2021, 2022, 2023, 2024, 2025], "Extended 2021–2025")
    associations = association_audit(
        [("Primary 2023–2025", primary_window), ("Extended 2021–2025", extended_window)],
        n_perm=args.n_perm,
        n_boot=args.n_boot,
        alpha=args.alpha,
        seed=args.seed,
    )
    comparison = compare_windows(primary_window, extended_window)

    write_outputs(
        output_dir,
        input_audit,
        historical_matches.sort_values(["year", "destination_original"]),
        historical,
        primary_window,
        extended_window,
        associations,
        comparison,
        primary_source,
    )

    primary_assoc = associations[
        (associations["analysis_window"] == "Primary 2023–2025")
        & (associations["sample"] == "Primary")
    ].iloc[0]
    extended_assoc = associations[
        (associations["analysis_window"] == "Extended 2021–2025")
        & (associations["sample"] == "Primary")
    ].iloc[0]
    retained = int(comparison["class_retained"].sum())
    changed = comparison.loc[~comparison["class_retained"], "destination"].tolist()

    print(f"2021 input : {visitor_2021}::{audit_2021['source_sheet']}")
    print(f"2022 input : {visitor_2022}::{audit_2022['source_sheet']}")
    print(f"Primary    : {primary_source}")
    print(f"Output     : {output_dir}")
    print()
    print(
        f"Primary 2023–2025: n={int(primary_assoc['n'])}; "
        f"rho={primary_assoc['rho']:.3f}; permutation p={primary_assoc['permutation_p']:.3f}; "
        f"bootstrap 95% CI [{primary_assoc['bootstrap_95CI_low']:.3f}, "
        f"{primary_assoc['bootstrap_95CI_high']:.3f}]."
    )
    print(
        f"Extended 2021–2025: n={int(extended_assoc['n'])}; "
        f"rho={extended_assoc['rho']:.3f}; permutation p={extended_assoc['permutation_p']:.3f}; "
        f"bootstrap 95% CI [{extended_assoc['bootstrap_95CI_low']:.3f}, "
        f"{extended_assoc['bootstrap_95CI_high']:.3f}]."
    )
    print(f"Class retained: {retained}/{len(comparison)} ({100 * retained / len(comparison):.1f}%).")
    print(f"Changed destinations: {', '.join(changed) if changed else 'None'}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, KeyError, ValueError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(2) from error
